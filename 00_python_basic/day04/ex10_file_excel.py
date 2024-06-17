# from 모듈 import 함수이름
from openpyxl import load_workbook

# 수식이 아닌 값만 받아온다.
wb = load_workbook("day04/data/excel01.xlsx" , data_only=True)

w_list = wb.sheetnames
print(w_list)

print("ㅡ" * 30)

# 타이틀을 이용해서 sheet 얻기
ws = wb['타이틀 연습']
print(ws['A1'].value) # Hello World
print(ws['A5'].value) # None (데이터가 없으면 오류가 아니라)

print("ㅡ" * 30)

# 범위 지정
get_cell = ws['A1' : 'C4']

for i in get_cell :
    for j in i :
        print(j.value, end= ' ')
    print()
print("ㅡ" * 30)

get_cell = ws['A1' : 'C4']

for i in get_cell :
    for j in i :
        if j.value == None : continue
        print(j.value, end= ' ')
    print()
print("ㅡ" * 30)

# 데이터를 리스트에 담기
t_list = []
get_cell = ws['A1' : 'C6']

for i in get_cell :
    for j in i :
        if j.value == None : continue
        t_list.append(j.value)

print(t_list)


print("ㅡ" * 30)