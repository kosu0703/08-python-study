# 엑셀 구조
# workbook - 엑셀파일 / worksheet - 엑셀시트 / cell - 시트 안에 존재하는 한칸 한칸

# 엑셀 관련 모듈
#   1) xlwt     2) OpenPyXL    3) pandas(빅데이터)

# openpyxl 설치 : 터미널에서 pip install openpyxl

# 1. 임포트 하기
# from 모듈 import 함수이름
from openpyxl import Workbook

# 2. 경로 지정
f_path = "day04/data/excel01.xlsx"

# 3. 엑셀 파일 만들기
# Workbook() 함수를 이용해서 workbook 객체 생성
# 생성시 기본적으로 sheet 하나가 생성된다.
wb = Workbook()
ws = wb.active

# 4. 시트 만들기
Sheet2 = wb.create_sheet("두번째 시트") # 추가
Sheet3 = wb.create_sheet("세번째 시트", 1) # 삽입

Sheet2.title = "성적관련정보"
ws.title = "타이틀 연습"

# 5. 셀에 데이터 넣기
ws['A1'] = "Hello World"
ws.cell(row=2, column=2, value="방가방가")
ws.cell(row=3, column=3).value="가나다라"
ws.cell(row=4, column=1).value="데이터 입력 연습"

# 범위 지정해서 두번째 시트에 데이터 넣기
for i in range(5) : 
    Sheet2.cell(row=1, column=i+1).value = "hi~~"
    
# 리스트 한번에 넣기
Sheet3.append([1, 2, 3, 4, 5, 6, 7, 8, 9])

# 시트 타이틀을 이용해서 데이터 입력
ws3 = wb['세번째 시트']
ws3['C1'] = "안녕하세요"

# 저장하고 파일 닫기
wb.save(f_path)
wb.close()

# 확장팩에서 Excel Viewer 설치하면 vscode 에서 엑셀파일 볼 수 있다.

